"""
Bulk import and export operations.
"""
import csv
import io
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from datetime import datetime

from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from accounts.models import User
from students.models import Student
from courses.models import Course, Class
from core.utils import StandardResponse
from accounts.permissions import IsAdmin, IsAdminOrRegistrar

bearer_security = [{'Bearer': []}]


class BulkOperationsViewSet(viewsets.GenericViewSet):
    """Bulk import and export operations."""
    
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_summary="Bulk Import Students",
        operation_description="Import students from CSV/Excel file",
        manual_parameters=[
            openapi.Parameter(
                'file',
                openapi.IN_FORM,
                type=openapi.TYPE_FILE,
                required=True,
                description='CSV or Excel file with student data'
            ),
            openapi.Parameter(
                'validate_only',
                openapi.IN_FORM,
                type=openapi.TYPE_BOOLEAN,
                default=False,
                description='Only validate, do not import'
            ),
        ],
        security=bearer_security
    )
    @action(
        detail=False,
        methods=['post'],
        url_path='students/import',
        parser_classes=[MultiPartParser],
        permission_classes=[IsAdminOrRegistrar]
    )
    def import_students(self, request):
        """
        Bulk import students from CSV/Excel.
        
        CSV Format:
        student_id,first_name,last_name,email,date_of_birth,gender,phone_number,address,city,state,postal_code,country
        """
        file = request.FILES.get('file')
        validate_only = request.data.get('validate_only', 'false').lower() == 'true'
        
        if not file:
            return Response(
                StandardResponse.error(
                    message='No file provided',
                    code='MISSING_FILE'
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Determine file type
        file_ext = file.name.split('.')[-1].lower()
        
        results = {
            'total_records': 0,
            'successful': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            if file_ext == 'csv':
                decoded_file = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                records = list(reader)
            elif file_ext in ['xlsx', 'xls']:
                wb = load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                records = [
                    {headers[i]: cell.value for i, cell in enumerate(row)}
                    for row in ws.iter_rows(min_row=2, values_only=False)
                ]
            else:
                return Response(
                    StandardResponse.error(
                        message='Unsupported file format. Use CSV or Excel',
                        code='INVALID_FILE_FORMAT'
                    ),
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results['total_records'] = len(records)
            
            for idx, record in enumerate(records, start=2):
                try:
                    # Extract data
                    student_id = str(record.get('student_id', '')).strip()
                    first_name = str(record.get('first_name', '')).strip()
                    last_name = str(record.get('last_name', '')).strip()
                    email = str(record.get('email', '')).strip()
                    
                    # Validation
                    if not all([student_id, first_name, last_name, email]):
                        results['errors'].append({
                            'row': idx,
                            'student_id': student_id,
                            'error': 'Missing required fields'
                        })
                        results['failed'] += 1
                        continue
                    
                    # Check for duplicates
                    if Student.objects.filter(student_id=student_id).exists():
                        results['errors'].append({
                            'row': idx,
                            'student_id': student_id,
                            'error': 'Duplicate student ID'
                        })
                        results['failed'] += 1
                        continue
                    
                    if User.objects.filter(email=email).exists():
                        results['errors'].append({
                            'row': idx,
                            'student_id': student_id,
                            'error': 'Email already exists'
                        })
                        results['failed'] += 1
                        continue
                    
                    # Create user and student (if not validate_only)
                    if not validate_only:
                        # Create user
                        username = email.split('@')[0]
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password='ChangeMe123!',  # Default password
                            first_name=first_name,
                            last_name=last_name,
                            role='STUDENT'
                        )
                        
                        # Create student profile
                        Student.objects.create(
                            user=user,
                            student_id=student_id,
                            date_of_birth=record.get('date_of_birth', '2000-01-01'),
                            gender=record.get('gender', 'OTHER'),
                            address=record.get('address', ''),
                            city=record.get('city', ''),
                            state=record.get('state', ''),
                            postal_code=record.get('postal_code', ''),
                            country=record.get('country', ''),
                            enrollment_date=datetime.now().date()
                        )
                    
                    results['successful'] += 1
                    
                except Exception as e:
                    results['errors'].append({
                        'row': idx,
                        'student_id': student_id if 'student_id' in locals() else 'N/A',
                        'error': str(e)
                    })
                    results['failed'] += 1
            
            message = 'Validation completed' if validate_only else 'Import completed'
            
            return Response(
                StandardResponse.success(
                    message=message,
                    data=results
                ),
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            return Response(
                StandardResponse.error(
                    message=f'Error processing file: {str(e)}',
                    code='PROCESSING_ERROR'
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @swagger_auto_schema(
        operation_summary="Export Students",
        manual_parameters=[
            openapi.Parameter('format', openapi.IN_QUERY, type=openapi.TYPE_STRING, 
                            enum=['csv', 'xlsx'], default='csv'),
            openapi.Parameter('academic_status', openapi.IN_QUERY, type=openapi.TYPE_STRING),
        ],
        security=bearer_security
    )
    @action(
        detail=False,
        methods=['get'],
        url_path='students/export',
        permission_classes=[IsAdminOrRegistrar]
    )
    def export_students(self, request):
        """Export students to CSV or Excel."""
        format_type = request.query_params.get('format', 'csv')
        academic_status = request.query_params.get('academic_status')
        
        # Query students
        queryset = Student.objects.select_related('user').all()
        
        if academic_status:
            queryset = queryset.filter(academic_status=academic_status)
        
        if format_type == 'csv':
            # Create CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Student ID', 'First Name', 'Last Name', 'Email',
                'Date of Birth', 'Gender', 'Phone', 'Address',
                'City', 'State', 'Postal Code', 'Country',
                'Enrollment Date', 'Academic Status', 'GPA'
            ])
            
            for student in queryset:
                writer.writerow([
                    student.student_id,
                    student.user.first_name,
                    student.user.last_name,
                    student.user.email,
                    student.date_of_birth,
                    student.gender,
                    student.user.phone_number or '',
                    student.address,
                    student.city,
                    student.state,
                    student.postal_code,
                    student.country,
                    student.enrollment_date,
                    student.get_academic_status_display(),
                    float(student.gpa)
                ])
            
            return response
            
        elif format_type == 'xlsx':
            # Create Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Headers
            headers = [
                'Student ID', 'First Name', 'Last Name', 'Email',
                'Date of Birth', 'Gender', 'Phone', 'Address',
                'City', 'State', 'Postal Code', 'Country',
                'Enrollment Date', 'Academic Status', 'GPA'
            ]
            ws.append(headers)
            
            # Data
            for student in queryset:
                ws.append([
                    student.student_id,
                    student.user.first_name,
                    student.user.last_name,
                    student.user.email,
                    student.date_of_birth,
                    student.gender,
                    student.user.phone_number or '',
                    student.address,
                    student.city,
                    student.state,
                    student.postal_code,
                    student.country,
                    student.enrollment_date,
                    student.get_academic_status_display(),
                    float(student.gpa)
                ])
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            
            return response
        
        return Response(
            StandardResponse.error(
                message='Invalid format. Use csv or xlsx',
                code='INVALID_FORMAT'
            ),
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @swagger_auto_schema(
        operation_summary="Export Courses",
        manual_parameters=[
            openapi.Parameter('format', openapi.IN_QUERY, type=openapi.TYPE_STRING, 
                            enum=['csv', 'xlsx'], default='csv'),
        ],
        security=bearer_security
    )
    @action(
        detail=False,
        methods=['get'],
        url_path='courses/export',
        permission_classes=[IsAdmin]
    )
    def export_courses(self, request):
        """Export courses to CSV or Excel."""
        format_type = request.query_params.get('format', 'csv')
        queryset = Course.objects.all()
        
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Course Code', 'Course Name', 'Description',
                'Credits', 'Department', 'Is Active'
            ])
            
            for course in queryset:
                writer.writerow([
                    course.course_code,
                    course.course_name,
                    course.description,
                    course.credits,
                    course.department,
                    course.is_active
                ])
            
            return response
        
        elif format_type == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Courses"
            
            headers = [
                'Course Code', 'Course Name', 'Description',
                'Credits', 'Department', 'Is Active'
            ]
            ws.append(headers)
            
            for course in queryset:
                ws.append([
                    course.course_code,
                    course.course_name,
                    course.description,
                    course.credits,
                    course.department,
                    course.is_active
                ])
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            
            return response
        
        return Response(
            StandardResponse.error(
                message='Invalid format',
                code='INVALID_FORMAT'
            ),
            status=status.HTTP_400_BAD_REQUEST
        )